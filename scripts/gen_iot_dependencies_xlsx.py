"""Generate IOT_DEPENDENCIES.xlsx — dependencies QUAD needs to support Qualcomm IoT SoCs."""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = Path(r"C:\work\05\QUAD\docs\IOT_DEPENDENCIES.xlsx")

HEADERS = [
    "ID",
    "Category",
    "Component",
    "Type",
    "Target Version",
    "Purpose (why QUAD needs it)",
    "IoT Capability Enabled",
    "Priority",
    "Status",
    "License",
    "Source / URL",
    "Notes",
]

# (Category, Component, Type, Version, Purpose, Capability, Priority, Status, License, URL, Notes)
ROWS: list[tuple] = [
    # Hardware / Qualcomm IoT SoCs
    ("Hardware / SoC", "Snapdragon QCS6490", "SoC", "—",
     "Reference IoT/robotics SoC for QUAD adapters", "Edge AI, Vision, Robotics", "P1", "Required",
     "Proprietary", "https://www.qualcomm.com/products/internet-of-things/industrial/processors/qcs6490",
     "Pairs with RB3 Gen 2 dev kit"),
    ("Hardware / SoC", "Snapdragon QCS8550", "SoC", "—",
     "High-perf IoT/AI target", "Edge AI, Industrial, Smart Camera", "P1", "Required",
     "Proprietary", "https://www.qualcomm.com/products/internet-of-things/industrial/processors/qcs8550",
     "12 TOPS NPU"),
    ("Hardware / SoC", "Snapdragon QCS610 / QCS605", "SoC", "—",
     "Camera + AI IoT class", "Smart Camera, AI Vision", "P2", "Required",
     "Proprietary", "https://www.qualcomm.com/products/internet-of-things/consumer/cameras/qcs610",
     "Hexagon DSP + NPU"),
    ("Hardware / SoC", "Snapdragon QCS8250 (RB5)", "SoC", "—",
     "Robotics/edge AI reference", "Robotics, Drones, Industrial AI", "P2", "Required",
     "Proprietary", "https://www.qualcomm.com/products/internet-of-things/robotics/qrb5165",
     "Used on RB5 dev kit"),
    ("Hardware / SoC", "Snapdragon QCM2290 / QCS2290", "SoC", "—",
     "Entry-tier IoT", "Wearables, Smart Speakers", "P3", "Optional",
     "Proprietary", "https://www.qualcomm.com/products/internet-of-things/consumer/processors/qcm2290",
     "Cellular IoT"),
    ("Hardware / SoC", "Snapdragon X75 5G Modem-RF", "Modem", "—",
     "5G/4G/NB-IoT/LTE-M cellular IoT", "Cellular IoT, NB-IoT, LTE-M", "P2", "Required",
     "Proprietary", "https://www.qualcomm.com/products/technology/modems/snapdragon-x75",
     "For connected IoT devices"),
    ("Hardware / SoC", "Qualcomm 9205 LTE Modem", "Modem", "—",
     "Low-power LTE-M / NB-IoT", "Asset Tracking, Smart Meters", "P3", "Optional",
     "Proprietary", "https://www.qualcomm.com/products/technology/modems/9205-lte",
     "Tiny battery-powered IoT"),
    ("Hardware / SoC", "Snapdragon Wear W5+", "SoC", "—",
     "Wearable IoT class", "Wearables, Hearables", "P3", "Optional",
     "Proprietary", "https://www.qualcomm.com/products/snapdragon-w5-plus-gen-1-wearable-platform",
     "Battery-optimized"),
    ("Hardware / Dev Board", "Qualcomm RB3 Gen 2", "Dev Board", "—",
     "Reference IoT dev kit (QCS6490)", "Hardware bring-up, CI lab", "P1", "Required",
     "Hardware", "https://www.qualcomm.com/developer/hardware/rb3-gen-2-development-kit",
     "Primary IoT lab device"),
    ("Hardware / Dev Board", "Qualcomm RB5 Dev Kit", "Dev Board", "—",
     "QCS8250 robotics platform", "Robotics, Drones", "P2", "Required",
     "Hardware", "https://www.qualcomm.com/developer/hardware/rb5-development-kit", ""),
    ("Hardware / Dev Board", "Thundercomm TurboX C865 SoM", "Dev Board", "—",
     "QCS8550 module for prototyping", "Edge AI prototyping", "P2", "Optional",
     "Hardware", "https://www.thundercomm.com/", ""),

    # OS / Firmware
    ("OS / Firmware", "Yocto Project (meta-qcom)", "Build System", "Scarthgap (5.0+)",
     "Custom Linux images for Qualcomm IoT SoCs", "Linux gateway image build", "P1", "Required",
     "MIT/Mixed", "https://www.yoctoproject.org/", "meta-qcom layer"),
    ("OS / Firmware", "Linux kernel", "Kernel", "6.1 LTS+",
     "Mainline kernel for IoT gateway-class", "Gateway/edge runtime", "P1", "Required",
     "GPL-2.0", "https://www.kernel.org/", "Qualcomm BSP backports"),
    ("OS / Firmware", "Qualcomm Linux", "Distribution", "1.x",
     "Vendor Linux distro for IoT SoCs", "Reference distro", "P1", "Required",
     "Proprietary", "https://www.qualcomm.com/developer/software/qualcomm-linux", ""),
    ("OS / Firmware", "Zephyr RTOS", "RTOS", "3.7 LTS",
     "RTOS target for MCU-class IoT companion cores", "MCU IoT, sensor nodes", "P2", "Required",
     "Apache-2.0", "https://www.zephyrproject.org/", ""),
    ("OS / Firmware", "FreeRTOS", "RTOS", "11.x",
     "Tiny RTOS for constrained IoT", "MCU IoT", "P2", "Optional",
     "MIT", "https://www.freertos.org/", ""),
    ("OS / Firmware", "ThreadX (Azure RTOS)", "RTOS", "6.x",
     "Real-time core option", "Real-time IoT", "P3", "Optional",
     "MIT", "https://github.com/azure-rtos/threadx", ""),
    ("OS / Firmware", "U-Boot", "Bootloader", "2024.04+",
     "Boot stack for IoT boards", "Device boot, OTA staging", "P1", "Required",
     "GPL-2.0", "https://u-boot.readthedocs.io/", ""),
    ("OS / Firmware", "Trusted Firmware-A (TF-A)", "Secure Boot", "2.10+",
     "ARMv8-A secure boot/EL3 firmware", "Secure boot chain", "P1", "Required",
     "BSD-3-Clause", "https://www.trustedfirmware.org/projects/tf-a/", ""),
    ("OS / Firmware", "OP-TEE", "TEE OS", "4.x",
     "Trusted Execution Environment", "Secure key/cert storage", "P2", "Required",
     "BSD-2-Clause", "https://www.op-tee.org/", "Pairs with QTEE on Qualcomm"),

    # Connectivity (Wireless)
    ("Connectivity", "BlueZ", "BLE Stack", "5.75+",
     "Linux Bluetooth stack for BLE-IoT", "BLE peripherals, beacons", "P1", "Required",
     "GPL-2.0", "http://www.bluez.org/", ""),
    ("Connectivity", "OpenThread", "Thread Stack", "Thread 1.4",
     "IPv6-mesh for low-power IoT", "Thread mesh, Matter substrate", "P1", "Required",
     "BSD-3-Clause", "https://openthread.io/", ""),
    ("Connectivity", "Matter (CHIP) SDK", "App Framework", "1.4",
     "Smart-home interop standard", "Smart home, Matter devices", "P1", "Required",
     "Apache-2.0", "https://github.com/project-chip/connectedhomeip", "CSA certification"),
    ("Connectivity", "ZBOSS Zigbee Stack", "Zigbee Stack", "3.x",
     "Zigbee 3.0 / Pro support", "Zigbee mesh", "P2", "Optional",
     "Proprietary/EULA", "https://dsr-iot.com/products/zboss-zigbee-3-software/", ""),
    ("Connectivity", "Z-Wave SDK (Silicon Labs)", "Z-Wave Stack", "7.x",
     "Z-Wave Plus support", "Z-Wave devices", "P3", "Optional",
     "Proprietary", "https://www.silabs.com/wireless/z-wave", ""),
    ("Connectivity", "ChirpStack / LoRa Basics", "LoRaWAN", "4.x",
     "LoRaWAN gateway/server stack", "LoRaWAN sensor networks", "P2", "Optional",
     "MIT/Apache-2.0", "https://www.chirpstack.io/", ""),
    ("Connectivity", "hostapd / wpa_supplicant", "Wi-Fi", "2.10+",
     "Wi-Fi STA/AP control plane", "Wi-Fi 6/7 connectivity", "P1", "Required",
     "BSD", "https://w1.fi/", ""),
    ("Connectivity", "ModemManager + libqmi + libmbim", "Cellular", "1.22+",
     "Cellular modem control on Linux", "Cellular IoT, NB-IoT, LTE-M", "P1", "Required",
     "GPL-2.0/LGPL", "https://modemmanager.org/", "Qualcomm QMI/MBIM"),
    ("Connectivity", "oFono", "Cellular", "2.x",
     "Alt cellular telephony stack", "Cellular IoT", "P3", "Optional",
     "GPL-2.0", "https://01.org/ofono", ""),
    ("Connectivity", "wpan-tools / nl802154", "802.15.4", "—",
     "Userspace 802.15.4 control", "Sub-GHz IoT, Thread", "P2", "Required",
     "GPL-2.0", "https://github.com/linux-wpan/wpan-tools", ""),

    # IoT Application Protocols
    ("Protocol", "Eclipse Mosquitto", "MQTT Broker", "2.0+",
     "Local MQTT broker for testing/edge", "MQTT pub/sub", "P1", "Required",
     "EPL-2.0/EDL", "https://mosquitto.org/", ""),
    ("Protocol", "paho-mqtt (Python)", "MQTT Client", "2.x",
     "Python MQTT client for QUAD adapters", "Telemetry, command pub/sub", "P1", "Required",
     "EPL-2.0/EDL", "https://pypi.org/project/paho-mqtt/", ""),
    ("Protocol", "MQTT-SN gateway", "MQTT-SN", "—",
     "MQTT for sensor networks (UDP)", "Constrained MQTT", "P3", "Optional",
     "EPL-2.0", "https://github.com/eclipse/paho.mqtt-sn.embedded-c", ""),
    ("Protocol", "libcoap", "CoAP", "4.3+",
     "C library CoAP client/server", "RESTful constrained IoT", "P1", "Required",
     "BSD-2-Clause", "https://libcoap.net/", ""),
    ("Protocol", "aiocoap (Python)", "CoAP", "0.4.x",
     "Async CoAP for QUAD bindings", "CoAP from Python adapters", "P1", "Required",
     "MIT", "https://aiocoap.readthedocs.io/", ""),
    ("Protocol", "Anjay LwM2M", "LwM2M", "3.x",
     "OMA LwM2M client/server", "Device mgmt, telemetry", "P2", "Required",
     "Apache-2.0", "https://avsystem.github.io/Anjay-doc/", ""),
    ("Protocol", "open62541", "OPC-UA", "1.4+",
     "Industrial OPC-UA stack", "Industrial IoT (IIoT)", "P2", "Optional",
     "MPL-2.0", "https://open62541.org/", ""),
    ("Protocol", "websockets / aiohttp", "WebSocket / HTTP", "12+ / 3.10+",
     "Async HTTP/WebSocket client+server", "Cloud uplink, REST APIs", "P1", "Required",
     "BSD/Apache-2.0", "https://pypi.org/project/websockets/", ""),
    ("Protocol", "grpcio", "gRPC", "1.60+",
     "gRPC for service-to-service IoT", "Edge microservices", "P2", "Optional",
     "Apache-2.0", "https://grpc.io/", ""),
    ("Protocol", "CBOR2 / msgpack", "Serialization", "5.x / 1.0+",
     "Compact binary encoding", "CoAP/LwM2M payloads", "P1", "Required",
     "MIT", "https://pypi.org/project/cbor2/", ""),
    ("Protocol", "protobuf", "Serialization", "5.x",
     "Protocol Buffers", "gRPC, telemetry schemas", "P2", "Required",
     "BSD-3-Clause", "https://protobuf.dev/", ""),

    # Cloud & Device Management
    ("Cloud / Device Mgmt", "AWS IoT Device SDK v2 (Python)", "Cloud SDK", "1.x",
     "AWS IoT Core uplink", "Cloud telemetry, jobs, shadow", "P1", "Required",
     "Apache-2.0", "https://github.com/aws/aws-iot-device-sdk-python-v2", ""),
    ("Cloud / Device Mgmt", "AWS IoT Greengrass v2", "Edge Runtime", "2.x",
     "Edge runtime + Lambda at edge", "Edge orchestration", "P2", "Optional",
     "Apache-2.0", "https://aws.amazon.com/greengrass/", ""),
    ("Cloud / Device Mgmt", "azure-iot-device (Python)", "Cloud SDK", "2.x",
     "Azure IoT Hub uplink", "Cloud telemetry, twin, methods", "P1", "Required",
     "MIT", "https://github.com/Azure/azure-iot-sdk-python", ""),
    ("Cloud / Device Mgmt", "Azure IoT Edge runtime", "Edge Runtime", "1.5+",
     "Containerized edge modules", "Edge ML modules", "P2", "Optional",
     "MIT", "https://github.com/Azure/iotedge", ""),
    ("Cloud / Device Mgmt", "Azure DPS / AWS Fleet Provisioning", "Provisioning", "—",
     "Zero-touch device onboarding", "Bulk provisioning", "P2", "Required",
     "Vendor", "https://learn.microsoft.com/azure/iot-dps/", ""),
    ("Cloud / Device Mgmt", "EdgeX Foundry", "Edge Platform", "3.x",
     "Vendor-neutral edge platform", "Industrial edge", "P3", "Optional",
     "Apache-2.0", "https://www.edgexfoundry.org/", ""),
    ("Cloud / Device Mgmt", "ThingsBoard", "IoT Platform", "3.7+",
     "Open-source IoT platform", "Self-hosted dashboards", "P3", "Optional",
     "Apache-2.0", "https://thingsboard.io/", ""),
    ("Cloud / Device Mgmt", "Mender", "OTA", "3.x",
     "Robust A/B OTA updates", "Fleet OTA", "P1", "Required",
     "Apache-2.0", "https://mender.io/", ""),
    ("Cloud / Device Mgmt", "RAUC", "OTA", "1.11+",
     "A/B image updates with rollback", "OTA fallback", "P2", "Optional",
     "LGPL-2.1", "https://rauc.io/", ""),
    ("Cloud / Device Mgmt", "SWUpdate", "OTA", "2024.x",
     "Yocto-friendly OTA", "Yocto-based OTA", "P2", "Optional",
     "GPL-2.0", "https://sbabic.github.io/swupdate/", ""),
    ("Cloud / Device Mgmt", "fwupd / LVFS", "Firmware Update", "1.9+",
     "Component firmware updates", "BIOS/firmware OTA", "P3", "Optional",
     "LGPL-2.1", "https://fwupd.org/", ""),

    # Security
    ("Security", "OpenSSL", "Crypto / TLS", "3.2+",
     "TLS 1.3, X.509, crypto primitives", "Secure transport", "P1", "Required",
     "Apache-2.0", "https://www.openssl.org/", ""),
    ("Security", "mbedTLS", "Crypto / TLS", "3.6 LTS",
     "Compact TLS for constrained devices", "MCU TLS / DTLS", "P1", "Required",
     "Apache-2.0", "https://www.trustedfirmware.org/projects/mbed-tls/", ""),
    ("Security", "wolfSSL", "Crypto / TLS", "5.7+",
     "Tiny TLS option", "MCU TLS alt", "P3", "Optional",
     "GPL-2.0/Commercial", "https://www.wolfssl.com/", ""),
    ("Security", "TF-M (Trusted Firmware-M)", "TEE (MCU)", "2.x",
     "PSA Crypto API on Cortex-M", "MCU root of trust", "P2", "Required",
     "BSD-3-Clause", "https://www.trustedfirmware.org/projects/tf-m/", ""),
    ("Security", "Qualcomm QTEE / SPU", "TEE", "—",
     "Hardware secure processing unit", "Key storage, secure boot", "P1", "Required",
     "Proprietary", "https://www.qualcomm.com/products/internet-of-things/security",
     "Per-SoC SDK"),
    ("Security", "PKCS#11 (p11-kit)", "HSM API", "0.25+",
     "Standard HSM/secure-element API", "Hardware-backed keys", "P2", "Required",
     "BSD-3-Clause", "https://p11-glue.github.io/p11-glue/p11-kit.html", ""),
    ("Security", "OSCORE (libcoap-oscore)", "App-layer Security", "—",
     "Object security for CoAP", "Constrained-IoT security", "P3", "Optional",
     "BSD-2-Clause", "https://datatracker.ietf.org/doc/html/rfc8613", ""),
    ("Security", "Matter Device Attestation", "Attestation", "1.4",
     "Cert-based device attestation", "Smart-home onboarding", "P2", "Required",
     "Apache-2.0", "https://csa-iot.org/all-solutions/matter/", ""),
    ("Security", "SBOM tooling (Syft / CycloneDX)", "Compliance", "Latest",
     "Generate SBOM per image", "Supply-chain security", "P2", "Required",
     "Apache-2.0", "https://github.com/anchore/syft", ""),

    # Edge AI Runtime (extending QUAD core to IoT)
    ("Edge AI Runtime", "QNN SDK / QAIRT", "Inference SDK", "2.x",
     "Primary NN runtime on Qualcomm IoT", "On-device inference", "P1", "Required",
     "Proprietary", "https://www.qualcomm.com/developer/software/qualcomm-ai-engine-direct-sdk",
     "Existing QUAD dependency"),
    ("Edge AI Runtime", "SNPE SDK", "Inference SDK", "2.x",
     "Legacy DLC runtime — useful for some IoT SoCs", "Inference on QCS6490/QCS8250", "P1", "Required",
     "Proprietary", "https://www.qualcomm.com/developer/software/neural-processing-sdk-for-ai", ""),
    ("Edge AI Runtime", "Hexagon SDK", "DSP/HTP SDK", "5.x",
     "DSP/NPU programming for IoT class", "HVX/HMX kernels", "P1", "Required",
     "Proprietary", "https://developer.qualcomm.com/software/hexagon-dsp-sdk", ""),
    ("Edge AI Runtime", "AIMET", "Quantization", "1.34+",
     "INT8/INT4 quantization", "Model compression for IoT", "P2", "Required",
     "BSD-3-Clause", "https://github.com/quic/aimet", ""),
    ("Edge AI Runtime", "Qualcomm AI Hub", "Cloud Profiling", "—",
     "Cloud benchmarking on Qualcomm devices", "CI benchmarking", "P2", "Optional",
     "Proprietary", "https://aihub.qualcomm.com/", ""),
    ("Edge AI Runtime", "ONNX Runtime", "Inference SDK", "1.18+",
     "Cross-platform inference fallback", "Reference / fallback path", "P2", "Optional",
     "MIT", "https://onnxruntime.ai/", "QNN EP available"),
    ("Edge AI Runtime", "TensorFlow Lite Micro", "Inference (MCU)", "head",
     "MCU inference path", "MCU AI", "P3", "Optional",
     "Apache-2.0", "https://github.com/tensorflow/tflite-micro", ""),
    ("Edge AI Runtime", "Apache TVM / microTVM", "Compiler", "0.16+",
     "Auto-tuned codegen", "Custom kernels for IoT", "P3", "Optional",
     "Apache-2.0", "https://tvm.apache.org/", ""),

    # Sensor / Hardware abstraction
    ("Sensors / HAL", "libgpiod (Python lgpio)", "GPIO", "2.1+ / 0.2+",
     "GPIO control on Linux", "Sensor IO", "P1", "Required",
     "LGPL-2.1 / LGPL-3.0", "https://git.kernel.org/pub/scm/libs/libgpiod/libgpiod.git/", ""),
    ("Sensors / HAL", "i2c-tools / smbus2", "I2C", "4.x / 0.4+",
     "I2C device access", "Sensor bus", "P1", "Required",
     "GPL-2.0 / MIT", "https://pypi.org/project/smbus2/", ""),
    ("Sensors / HAL", "spidev (Python)", "SPI", "3.x",
     "SPI access from Python", "High-speed sensors", "P1", "Required",
     "GPL-2.0", "https://pypi.org/project/spidev/", ""),
    ("Sensors / HAL", "pyserial", "UART", "3.5+",
     "Serial/UART for legacy IoT", "Modem AT, sensor UARTs", "P1", "Required",
     "BSD-3-Clause", "https://pypi.org/project/pyserial/", ""),
    ("Sensors / HAL", "python-can + SocketCAN", "CAN", "4.x",
     "CAN bus access (auto/industrial)", "Vehicular & industrial IoT", "P2", "Optional",
     "LGPL-3.0", "https://python-can.readthedocs.io/", ""),
    ("Sensors / HAL", "pymodbus", "Modbus", "3.7+",
     "Modbus RTU/TCP", "Industrial sensors / PLC", "P2", "Optional",
     "BSD-3-Clause", "https://pymodbus.readthedocs.io/", ""),
    ("Sensors / HAL", "Linux IIO subsystem", "IIO", "kernel",
     "Industrial I/O sensors", "ADC, IMU, env sensors", "P2", "Required",
     "GPL-2.0", "https://www.kernel.org/doc/html/latest/driver-api/iio/index.html", ""),
    ("Sensors / HAL", "bleak (Python BLE)", "BLE Client", "0.22+",
     "Cross-platform BLE central", "BLE sensor onboarding", "P2", "Required",
     "MIT", "https://bleak.readthedocs.io/", ""),

    # Containers / Edge orchestration
    ("Edge Orchestration", "Docker Engine", "Container Runtime", "26+",
     "Edge containerization", "Containerized adapters", "P1", "Required",
     "Apache-2.0", "https://docs.docker.com/engine/", ""),
    ("Edge Orchestration", "Podman", "Container Runtime", "5.x",
     "Rootless container alt", "Edge containers (rootless)", "P3", "Optional",
     "Apache-2.0", "https://podman.io/", ""),
    ("Edge Orchestration", "K3s", "Kubernetes", "1.30+",
     "Lightweight K8s for IoT gateways", "Multi-node fleet", "P2", "Optional",
     "Apache-2.0", "https://k3s.io/", ""),
    ("Edge Orchestration", "KubeEdge", "Kubernetes", "1.18+",
     "Edge-native K8s", "Cloud-edge sync", "P3", "Optional",
     "Apache-2.0", "https://kubeedge.io/", ""),
    ("Edge Orchestration", "wasmtime / WasmEdge", "WASM Runtime", "Latest",
     "Sandboxed WASM modules at edge", "Portable edge functions", "P3", "Optional",
     "Apache-2.0", "https://wasmtime.dev/", ""),

    # Telemetry / Observability
    ("Telemetry", "OpenTelemetry SDK (Python/C++)", "Tracing/Metrics", "1.27+",
     "Standard observability", "Edge tracing & metrics", "P1", "Required",
     "Apache-2.0", "https://opentelemetry.io/", ""),
    ("Telemetry", "Prometheus client_python", "Metrics", "0.20+",
     "Prom metrics from QUAD adapters", "Edge metrics scrape", "P2", "Required",
     "Apache-2.0", "https://github.com/prometheus/client_python", ""),
    ("Telemetry", "Fluent Bit", "Log Forwarder", "3.x",
     "Edge log shipping", "Log aggregation", "P2", "Optional",
     "Apache-2.0", "https://fluentbit.io/", ""),
    ("Telemetry", "InfluxDB / Telegraf", "Time-series DB", "2.7+ / 1.30+",
     "Telemetry storage", "Local time-series buffer", "P3", "Optional",
     "MIT", "https://www.influxdata.com/", ""),

    # Build / Cross-compile / Tooling
    ("Build / Tooling", "CMake", "Build Tool", "3.25+",
     "C/C++ build for adapters", "Native code build", "P1", "Required",
     "BSD-3-Clause", "https://cmake.org/", ""),
    ("Build / Tooling", "ARM GCC Cross Toolchain", "Toolchain", "13.2+",
     "Cross-compile for ARM IoT", "Cross-build", "P1", "Required",
     "GPL-3.0", "https://developer.arm.com/Tools and Software/GNU Toolchain", ""),
    ("Build / Tooling", "LLVM / Clang", "Toolchain", "17+",
     "Hexagon backend & general LLVM", "DSP kernels", "P1", "Required",
     "Apache-2.0", "https://llvm.org/", ""),
    ("Build / Tooling", "BitBake / OpenEmbedded", "Build System", "Scarthgap",
     "Yocto build engine", "Image assembly", "P1", "Required",
     "GPL-2.0", "https://docs.yoctoproject.org/", ""),
    ("Build / Tooling", "Buildroot", "Build System", "2024.02+",
     "Alt embedded Linux builder", "Lean Linux images", "P3", "Optional",
     "GPL-2.0", "https://buildroot.org/", ""),
    ("Build / Tooling", "QEMU", "Emulator", "9.x",
     "ARM64/MIPS emulation for CI", "Pre-silicon CI", "P2", "Required",
     "GPL-2.0", "https://www.qemu.org/", ""),
    ("Build / Tooling", "Renode", "Simulator", "1.15+",
     "MCU full-system simulation", "MCU CI", "P3", "Optional",
     "MIT", "https://renode.io/", ""),
    ("Build / Tooling", "ccache / sccache", "Build Cache", "Latest",
     "Speed up cross-compiles", "CI build cache", "P2", "Optional",
     "BSD/Apache-2.0", "https://ccache.dev/", ""),

    # Python adapter dependencies for QUAD
    ("Python (QUAD adapters)", "fastmcp", "MCP server", "Latest",
     "Existing QUAD MCP framework", "Expose IoT tools via MCP", "P1", "Required",
     "Apache-2.0", "https://github.com/jlowin/fastmcp", "Already in QUAD"),
    ("Python (QUAD adapters)", "pydantic", "Validation", "2.7+",
     "Schema validation for IoT models", "Tool I/O schemas", "P1", "Required",
     "MIT", "https://docs.pydantic.dev/", "Already in QUAD"),
    ("Python (QUAD adapters)", "httpx", "HTTP client", "0.27+",
     "Cloud APIs / IoT REST", "REST uplink", "P1", "Required",
     "BSD-3-Clause", "https://www.python-httpx.org/", "Already in QUAD"),
    ("Python (QUAD adapters)", "asyncio-mqtt / gmqtt", "MQTT (async)", "0.16+ / 0.7+",
     "Async MQTT for IoT adapters", "Telemetry pub/sub", "P2", "Optional",
     "BSD-3-Clause", "https://github.com/sbtinstruments/asyncio-mqtt", ""),
    ("Python (QUAD adapters)", "cryptography", "Crypto", "43+",
     "X.509, PKCS, JWT", "Device identity", "P1", "Required",
     "Apache-2.0/BSD", "https://cryptography.io/", ""),
    ("Python (QUAD adapters)", "PyJWT", "JWT", "2.9+",
     "Cloud auth tokens", "Cloud auth", "P2", "Required",
     "MIT", "https://pyjwt.readthedocs.io/", ""),
    ("Python (QUAD adapters)", "psutil", "System info", "5.9+",
     "Detect IoT host resources", "hardware_detect on IoT", "P1", "Required",
     "BSD-3-Clause", "https://psutil.readthedocs.io/", ""),
    ("Python (QUAD adapters)", "fabric / paramiko", "SSH", "3.x / 3.4+",
     "Remote control of IoT boards", "Bring-up via SSH", "P1", "Required",
     "BSD-2-Clause / LGPL-2.1", "https://www.fabfile.org/", ""),

    # Standards / Certification
    ("Compliance / Cert", "CSA Matter Certification", "Certification", "1.4",
     "Required for Matter-branded devices", "Smart-home shipping", "P2", "Required",
     "Program", "https://csa-iot.org/all-solutions/matter/", "Per-product fee"),
    ("Compliance / Cert", "Bluetooth SIG Qualification", "Certification", "—",
     "Required for BT-branded devices", "BLE shipping", "P2", "Required",
     "Program", "https://www.bluetooth.com/develop-with-bluetooth/qualification-listing/", ""),
    ("Compliance / Cert", "Wi-Fi Alliance Certification", "Certification", "—",
     "Wi-Fi 6/7 branding", "Wi-Fi shipping", "P2", "Required",
     "Program", "https://www.wi-fi.org/certification", ""),
    ("Compliance / Cert", "Thread Group Certification", "Certification", "1.4",
     "Required for Thread devices", "Thread shipping", "P3", "Optional",
     "Program", "https://www.threadgroup.org/", ""),
    ("Compliance / Cert", "FCC / CE / IC RF compliance", "Certification", "—",
     "Regulatory radio compliance", "Geo shipping", "P1", "Required",
     "Program", "https://www.fcc.gov/", "Per-region"),

    # CI / Test
    ("CI / Test", "pytest + pytest-asyncio", "Test framework", "8.x / 0.23+",
     "QUAD test stack", "Unit/integration tests", "P1", "Required",
     "MIT", "https://docs.pytest.org/", "Already in QUAD"),
    ("CI / Test", "AWS IoT Device Simulator", "Simulator", "—",
     "Synth fleet for cloud testing", "Cloud E2E tests", "P3", "Optional",
     "Apache-2.0", "https://aws.amazon.com/solutions/implementations/iot-device-simulator/", ""),
    ("CI / Test", "GitHub Actions self-hosted runner (ARM)", "CI infra", "Latest",
     "Run CI on Qualcomm ARM hardware", "Hardware CI", "P2", "Required",
     "MIT", "https://docs.github.com/actions/hosting-your-own-runners", ""),
    ("CI / Test", "robot-framework / pytest-iot fixtures", "E2E", "7.x",
     "Device-in-the-loop tests", "Hardware E2E", "P3", "Optional",
     "Apache-2.0", "https://robotframework.org/", ""),
]


def main() -> None:
    wb = Workbook()

    # ---- Overview sheet ----
    ws_o = wb.active
    ws_o.title = "Overview"

    title_font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="1F4E79")
    label_font = Font(name="Calibri", size=11, bold=True)
    body_font = Font(name="Calibri", size=11)

    ws_o["A1"] = "QUAD — IoT Device Support Dependencies"
    ws_o["A1"].font = title_font
    ws_o["A1"].fill = title_fill
    ws_o.merge_cells("A1:D1")
    ws_o.row_dimensions[1].height = 28

    meta = [
        ("Document",        "IOT_DEPENDENCIES.xlsx"),
        ("Owner",           "QUAD platform team"),
        ("Generated",       date.today().isoformat()),
        ("Scope",           "Qualcomm IoT SoCs (broad): RB3 Gen 2 / RB5, QCS6490 / QCS8550 / QCS610, "
                            "Linux gateway + RTOS companion, BLE / Wi-Fi / Thread / Matter, "
                            "MQTT / CoAP / LwM2M, AWS / Azure IoT, OTA, edge AI inference."),
        ("Companion docs",  "docs/PREREQUISITES.md, docs/CLAUDE.md, docs/IMPLEMENTATION_GUIDE.md"),
        ("Total entries",   str(len(ROWS))),
    ]

    r = 3
    for label, value in meta:
        ws_o.cell(row=r, column=1, value=label).font = label_font
        c = ws_o.cell(row=r, column=2, value=value)
        c.font = body_font
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws_o.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1

    r += 1
    ws_o.cell(row=r, column=1, value="Legend").font = title_font
    ws_o.cell(row=r, column=1).fill = title_fill
    ws_o.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws_o.row_dimensions[r].height = 22
    r += 1

    legend = [
        ("Priority",  "P1 = must-have for first IoT release; P2 = next milestone; P3 = stretch / optional."),
        ("Status",    "Required = needed for shipping the IoT scope; Optional = enables extra capabilities."),
        ("Type",      "SoC, Dev Board, SDK, Library, Service, Protocol, Tool, Certification."),
        ("Category",  "Hardware/SoC, OS/Firmware, Connectivity, Protocol, Cloud / Device Mgmt, Security, "
                      "Edge AI Runtime, Sensors / HAL, Edge Orchestration, Telemetry, Build / Tooling, "
                      "Python (QUAD adapters), Compliance / Cert, CI / Test."),
    ]
    for label, value in legend:
        ws_o.cell(row=r, column=1, value=label).font = label_font
        c = ws_o.cell(row=r, column=2, value=value)
        c.font = body_font
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws_o.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws_o.row_dimensions[r].height = 32
        r += 1

    ws_o.column_dimensions["A"].width = 18
    ws_o.column_dimensions["B"].width = 40
    ws_o.column_dimensions["C"].width = 30
    ws_o.column_dimensions["D"].width = 30

    # ---- Dependencies sheet ----
    ws = wb.create_sheet("Dependencies")

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws.append(HEADERS)
    for col_idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 30

    # Color map for category banding
    category_fill = {
        "Hardware / SoC":          PatternFill("solid", fgColor="FFF2CC"),
        "Hardware / Dev Board":    PatternFill("solid", fgColor="FFF2CC"),
        "OS / Firmware":           PatternFill("solid", fgColor="DEEBF7"),
        "Connectivity":            PatternFill("solid", fgColor="E2EFDA"),
        "Protocol":                PatternFill("solid", fgColor="FCE4D6"),
        "Cloud / Device Mgmt":     PatternFill("solid", fgColor="EDEDED"),
        "Security":                PatternFill("solid", fgColor="F8CBAD"),
        "Edge AI Runtime":         PatternFill("solid", fgColor="D9E1F2"),
        "Sensors / HAL":           PatternFill("solid", fgColor="FFF2CC"),
        "Edge Orchestration":      PatternFill("solid", fgColor="DEEBF7"),
        "Telemetry":               PatternFill("solid", fgColor="E2EFDA"),
        "Build / Tooling":         PatternFill("solid", fgColor="FCE4D6"),
        "Python (QUAD adapters)":  PatternFill("solid", fgColor="EDEDED"),
        "Compliance / Cert":       PatternFill("solid", fgColor="F8CBAD"),
        "CI / Test":               PatternFill("solid", fgColor="D9E1F2"),
    }

    priority_fill = {
        "P1": PatternFill("solid", fgColor="C6EFCE"),
        "P2": PatternFill("solid", fgColor="FFEB9C"),
        "P3": PatternFill("solid", fgColor="F4CCCC"),
    }

    for i, row in enumerate(ROWS, start=1):
        cat = row[0]
        ws.append([f"IOT-{i:03d}", *row])
        excel_row = i + 1
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=excel_row, column=col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            cell.font = body_font
            if cat in category_fill:
                cell.fill = category_fill[cat]
        prio_cell = ws.cell(row=excel_row, column=8)
        if prio_cell.value in priority_fill:
            prio_cell.fill = priority_fill[prio_cell.value]
            prio_cell.font = Font(name="Calibri", size=11, bold=True)
            prio_cell.alignment = Alignment(horizontal="center", vertical="center")
        url_cell = ws.cell(row=excel_row, column=11)
        if url_cell.value and isinstance(url_cell.value, str) and url_cell.value.startswith("http"):
            url_cell.hyperlink = url_cell.value
            url_cell.font = Font(name="Calibri", size=11, color="0563C1", underline="single")

    widths = {
        "A": 10,   # ID
        "B": 22,   # Category
        "C": 30,   # Component
        "D": 18,   # Type
        "E": 14,   # Version
        "F": 44,   # Purpose
        "G": 32,   # Capability
        "H": 10,   # Priority
        "I": 12,   # Status
        "J": 18,   # License
        "K": 50,   # URL
        "L": 30,   # Notes
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"

    # Add as Excel table for sort/filter
    last_col = get_column_letter(len(HEADERS))
    last_row = len(ROWS) + 1
    table = Table(displayName="IoTDependencies", ref=f"A1:{last_col}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    ws.add_table(table)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT} ({len(ROWS)} rows)")


if __name__ == "__main__":
    main()
